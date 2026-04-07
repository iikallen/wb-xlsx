from __future__ import annotations

import argparse
import concurrent.futures
import json
import math
import random
import threading
import time
from pathlib import Path
from typing import Any
from urllib.parse import quote_plus, urlencode

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from playwright.sync_api import sync_playwright


DEFAULT_QUERY = "пальто из натуральной шерсти"
DEFAULT_DEST = "-1257786"
DEFAULT_OUTPUT_DIR = Path.cwd() / "output"
DEFAULT_BATCH_SIZE = 40
DEFAULT_CARD_WORKERS = 12
DEFAULT_TIMEOUT_SECONDS = 60
DEFAULT_HEADLESS = False

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/135.0.0.0 Safari/537.36 Edg/135.0.0.0"
)

EDGE_CANDIDATES = [
    Path(r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"),
    Path(r"C:\Program Files\Microsoft\Edge\Application\msedge.exe"),
    Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe"),
]

_thread_local = threading.local()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Wildberries parser with XLSX export",
    )
    parser.add_argument("--query", default=DEFAULT_QUERY, help="Search query")
    parser.add_argument("--dest", default=DEFAULT_DEST, help="Destination code")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="Output directory")
    parser.add_argument("--batch-size", type=int, default=DEFAULT_BATCH_SIZE, help="Detail batch size")
    parser.add_argument(
        "--card-workers",
        type=int,
        default=DEFAULT_CARD_WORKERS,
        help="Concurrent workers for public card.json requests",
    )
    parser.add_argument("--limit", type=int, default=0, help="Limit number of products")
    parser.add_argument(
        "--timeout-seconds",
        type=int,
        default=DEFAULT_TIMEOUT_SECONDS,
        help="HTTP timeout",
    )
    parser.add_argument(
        "--headless",
        default=str(DEFAULT_HEADLESS).lower(),
        help="true/false for browser visibility",
    )
    return parser.parse_args()


def parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    return normalized in {"1", "true", "yes", "y", "on"}


def now_ms() -> int:
    return int(time.time() * 1000)


def normalize_text(value: Any) -> str:
    return " ".join(str(value or "").split())


def ordered_unique(values: list[int]) -> list[int]:
    seen: set[int] = set()
    result: list[int] = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def chunked(items: list[int], size: int) -> list[list[int]]:
    return [items[index : index + size] for index in range(0, len(items), size)]


def format_duration(seconds: float) -> str:
    rounded = int(round(seconds))
    minutes, secs = divmod(rounded, 60)
    if minutes:
        return f"{minutes}m {secs}s"
    return f"{secs}s"


def detect_browser_executable() -> Path:
    for candidate in EDGE_CANDIDATES:
        if candidate.exists():
            return candidate
    local_chrome = Path.home() / "AppData" / "Local" / "Google" / "Chrome" / "Application" / "chrome.exe"
    if local_chrome.exists():
        return local_chrome
    raise FileNotFoundError("Edge/Chrome executable not found")


def build_search_referer(query: str) -> str:
    return f"https://www.wildberries.ru/catalog/0/search.aspx?search={quote_plus(query)}"


def volume_from_nm_id(nm_id: int) -> int:
    return int(nm_id) // 100000


def part_from_nm_id(nm_id: int) -> int:
    return int(nm_id) // 1000


def build_product_url(nm_id: int) -> str:
    return f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx"


def build_seller_url(supplier_id: int | None) -> str:
    return f"https://www.wildberries.ru/seller/{supplier_id}" if supplier_id else ""


def get_thread_session(timeout_seconds: int) -> requests.Session:
    session = getattr(_thread_local, "session", None)
    if session is None:
        session = requests.Session()
        session.headers.update(
            {
                "User-Agent": USER_AGENT,
                "Accept": "application/json, text/plain, */*",
            }
        )
        _thread_local.session = session
    return session


def request_public_json(
    url: str,
    timeout_seconds: int,
    *,
    description: str,
    allow_404: bool = False,
) -> Any:
    last_error: Exception | None = None
    for attempt in range(1, 4):
        try:
            response = get_thread_session(timeout_seconds).get(url, timeout=timeout_seconds)
            if 200 <= response.status_code < 300:
                return response.json()
            if allow_404 and response.status_code == 404:
                return None
            last_error = RuntimeError(
                f"{description} failed with status {response.status_code} for {url}: {response.text[:180]}"
            )
        except Exception as error:  # noqa: BLE001
            last_error = error
        time.sleep(attempt)
    raise RuntimeError(f"{description} failed: {last_error}") from last_error


class WildberriesPythonClient:
    def __init__(
        self,
        *,
        query: str,
        dest: str,
        headless: bool,
        timeout_seconds: int,
    ) -> None:
        self.query = query
        self.dest = dest
        self.headless = headless
        self.timeout_seconds = timeout_seconds
        self.browser_path = detect_browser_executable()

        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": USER_AGENT,
                "Accept": "application/json, text/plain, */*",
            }
        )
        self.media_route_hosts: list[dict[str, Any]] = []

    def start(self) -> None:
        self.playwright = sync_playwright().start()
        self.browser = self.playwright.chromium.launch(
            headless=self.headless,
            executable_path=str(self.browser_path),
            args=["--disable-blink-features=AutomationControlled"],
        )
        self.context = self.browser.new_context(
            user_agent=USER_AGENT,
            locale="ru-RU",
            timezone_id="Europe/Moscow",
            viewport={"width": 1366, "height": 900},
            color_scheme="light",
        )
        self.context.add_init_script(
            """
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            Object.defineProperty(navigator, 'platform', { get: () => 'Win32' });
            Object.defineProperty(navigator, 'language', { get: () => 'ru-RU' });
            Object.defineProperty(navigator, 'languages', { get: () => ['ru-RU', 'ru', 'en-US', 'en'] });
            """
        )
        self.page = self.context.new_page()
        self.ensure_session()
        self.sync_cookies_to_session()
        self.media_route_hosts = self.fetch_media_route_hosts()

    def close(self) -> None:
        if self.browser is not None:
            self.browser.close()
        if self.playwright is not None:
            self.playwright.stop()
        self.session.close()

    def ensure_session(self) -> None:
        assert self.page is not None
        self.page.goto("https://www.wildberries.ru/", wait_until="domcontentloaded", timeout=self.timeout_seconds * 1000)
        started_at = time.time()
        while time.time() - started_at < 120:
            title = self.page.title()
            body_text = ""
            try:
                body_text = self.page.locator("body").inner_text(timeout=1500)
            except Exception:  # noqa: BLE001
                body_text = ""

            blocked = (
                "Почти готово" in title
                or "Подозрительная активность" in body_text
                or "Проверяем браузер" in body_text
                or "Что-то не так" in body_text
            )
            if not blocked:
                return

            print("[session] Wildberries показывает антибот-страницу, жду прохождения проверки...")
            self.page.wait_for_timeout(5000)

        raise RuntimeError(
            "Не удалось пройти антибот-проверку Wildberries за 120 секунд. Перезапустите скрипт."
        )

    def sync_cookies_to_session(self) -> None:
        assert self.context is not None
        self.session.cookies.clear()
        for cookie in self.context.cookies():
            self.session.cookies.set(
                cookie["name"],
                cookie["value"],
                domain=cookie.get("domain"),
                path=cookie.get("path"),
            )

    def refresh_session(self) -> None:
        self.ensure_session()
        self.sync_cookies_to_session()

    def request_json(
        self,
        url: str,
        *,
        description: str,
        headers: dict[str, str] | None = None,
        allow_404: bool = False,
    ) -> Any:
        last_error: Exception | None = None
        for attempt in range(1, 4):
            try:
                response = self.session.get(
                    url,
                    headers=headers,
                    timeout=self.timeout_seconds,
                )
                if 200 <= response.status_code < 300:
                    try:
                        return response.json()
                    except Exception as error:  # noqa: BLE001
                        raise RuntimeError(
                            f"{description} returned non-JSON response for {url}: {response.text[:180]}"
                        ) from error
                if allow_404 and response.status_code == 404:
                    return None
                if response.status_code in {403, 498} and attempt < 3:
                    print(f"[retry] {description}: статус {response.status_code}, обновляю браузерную сессию.")
                    self.refresh_session()
                    continue
                last_error = RuntimeError(
                    f"{description} failed with status {response.status_code} for {url}: {response.text[:180]}"
                )
            except Exception as error:  # noqa: BLE001
                last_error = error
                if attempt < 3:
                    time.sleep(attempt)
                    continue
            if attempt < 3:
                time.sleep(attempt)
        raise RuntimeError(f"{description} failed: {last_error}") from last_error

    def build_search_url(self, page_number: int) -> str:
        params = {
            "ab_testing": "false",
            "appType": "1",
            "curr": "rub",
            "dest": self.dest,
            "hide_dflags": "131072",
            "hide_dtype": "13",
            "hide_vflags": "4294967296",
            "inheritFilters": "false",
            "lang": "ru",
            "query": self.query,
            "resultset": "catalog",
            "sort": "popular",
            "spp": "30",
            "suppressSpellcheck": "false",
        }
        if page_number > 1:
            params["page"] = str(page_number)
        return "https://www.wildberries.ru/__internal/u-search/exactmatch/sng/common/v18/search?" + urlencode(params)

    def search_headers(self) -> dict[str, str]:
        return {
            "Referer": build_search_referer(self.query),
            "x-queryid": f"qid{now_ms()}{random.randint(1000, 9999)}",
            "x-userid": "0",
        }

    def fetch_search_page(self, page_number: int) -> dict[str, Any]:
        return self.request_json(
            self.build_search_url(page_number),
            description=f"search page {page_number}",
            headers=self.search_headers(),
        )

    def build_detail_url(self, product_ids: list[int]) -> str:
        params = {
            "appType": "1",
            "curr": "rub",
            "dest": self.dest,
            "spp": "30",
            "hide_vflags": "4294967296",
            "hide_dflags": "131072",
            "hide_dtype": "13",
            "ab_testing": "false",
            "lang": "ru",
            "nm": ";".join(str(product_id) for product_id in product_ids),
        }
        return "https://www.wildberries.ru/__internal/u-card/cards/v4/detail?" + urlencode(params)

    def fetch_detail_batch(self, product_ids: list[int]) -> dict[str, Any]:
        return self.request_json(
            self.build_detail_url(product_ids),
            description=f"detail batch {product_ids[0]}",
            headers={"Referer": build_search_referer(self.query)},
        )

    def fetch_media_route_hosts(self) -> list[dict[str, Any]]:
        data = request_public_json(
            f"https://cdn.wbbasket.ru/api/v3/upstreams?t={now_ms()}",
            self.timeout_seconds,
            description="cdn upstreams",
        )
        hosts = data.get("origin", {}).get("mediabasket_route_map", [{}])[0].get("hosts", [])
        if not hosts:
            raise RuntimeError("Не удалось получить карту basket-хостов.")
        return hosts

    def resolve_media_host(self, nm_id: int) -> str:
        volume = volume_from_nm_id(nm_id)
        for host in self.media_route_hosts:
            if host["vol_range_from"] <= volume <= host["vol_range_to"]:
                return host["host"]
        raise RuntimeError(f"Не найден basket-хост для товара {nm_id} (vol={volume}).")

    def build_card_info_url(self, nm_id: int) -> str:
        host = self.resolve_media_host(nm_id)
        return (
            f"https://{host}/vol{volume_from_nm_id(nm_id)}/part{part_from_nm_id(nm_id)}/"
            f"{nm_id}/info/ru/card.json"
        )

    def build_image_urls(self, nm_id: int, photo_count: int) -> list[str]:
        if photo_count <= 0:
            return []
        host = self.resolve_media_host(nm_id)
        return [
            (
                f"https://{host}/vol{volume_from_nm_id(nm_id)}/part{part_from_nm_id(nm_id)}/"
                f"{nm_id}/images/big/{image_index}.webp"
            )
            for image_index in range(1, photo_count + 1)
        ]

    def fetch_card_info(self, nm_id: int) -> dict[str, Any] | None:
        return request_public_json(
            self.build_card_info_url(nm_id),
            self.timeout_seconds,
            description=f"card.json {nm_id}",
            allow_404=True,
        )


def collect_product_ids(client: WildberriesPythonClient, limit: int) -> list[int]:
    first_page = client.fetch_search_page(1)
    first_products = first_page.get("products", [])
    total = int(first_page.get("total", 0))
    page_size = len(first_products) or 100
    total_pages = math.ceil(total / page_size) if page_size else 0

    product_ids: list[int] = [int(product["id"]) for product in first_products]
    print(f"[search] страница 1/{total_pages}, товаров найдено: {total}.")

    for page_number in range(2, total_pages + 1):
        if limit > 0 and len(ordered_unique(product_ids)) >= limit:
            break
        page_data = client.fetch_search_page(page_number)
        product_ids.extend(int(product["id"]) for product in page_data.get("products", []))
        print(f"[search] страница {page_number}/{total_pages}, накоплено товаров: {len(product_ids)}.")

    unique_ids = ordered_unique(product_ids)
    if limit > 0:
        unique_ids = unique_ids[:limit]
    return unique_ids


def collect_detail_map(
    client: WildberriesPythonClient,
    product_ids: list[int],
    batch_size: int,
) -> dict[int, dict[str, Any]]:
    detail_map: dict[int, dict[str, Any]] = {}
    batches = chunked(product_ids, batch_size)
    for index, batch in enumerate(batches, start=1):
        response = client.fetch_detail_batch(batch)
        for product in response.get("products", []):
            detail_map[int(product["id"])] = product
        print(f"[detail] батч {index}/{len(batches)}, получено карточек: {len(detail_map)}.")
    return detail_map


def collect_card_info_map(
    client: WildberriesPythonClient,
    product_ids: list[int],
    workers: int,
) -> dict[int, dict[str, Any] | None]:
    card_info_map: dict[int, dict[str, Any] | None] = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=max(1, workers)) as executor:
        future_map = {
            executor.submit(client.fetch_card_info, product_id): product_id for product_id in product_ids
        }
        total = len(future_map)
        for index, future in enumerate(concurrent.futures.as_completed(future_map), start=1):
            product_id = future_map[future]
            card_info_map[product_id] = future.result()
            if index % 100 == 0 or index == total:
                print(f"[card.json] обработано {index}/{total}.")
    return card_info_map


def grouped_options(card_info: dict[str, Any] | None) -> list[dict[str, Any]]:
    if not card_info:
        return []
    if card_info.get("grouped_options"):
        return card_info["grouped_options"]
    if card_info.get("options"):
        return [{"group_name": "Характеристики", "options": card_info["options"]}]
    return []


def find_characteristic(groups: list[dict[str, Any]], name: str) -> str:
    expected = normalize_text(name).lower()
    for group in groups:
        for option in group.get("options", []):
            option_name = normalize_text(option.get("name")).lower()
            if option_name == expected:
                return normalize_text(option.get("value"))
    return ""


def get_price(product: dict[str, Any]) -> float | None:
    prices: list[float] = []
    for size in product.get("sizes", []):
        raw_price = size.get("price", {}).get("product")
        if isinstance(raw_price, (int, float)) and raw_price > 0:
            prices.append(round(raw_price / 100, 2))
    return min(prices) if prices else None


def get_sizes(product: dict[str, Any]) -> list[str]:
    sizes: list[str] = []
    for size in product.get("sizes", []):
        name = normalize_text(size.get("origName") or size.get("name"))
        if name:
            sizes.append(name)
    return ordered_unique(sizes)


def get_total_stock(product: dict[str, Any]) -> int:
    total = 0
    for size in product.get("sizes", []):
        for stock in size.get("stocks", []):
            total += int(stock.get("qty") or 0)
    if total > 0:
        return total
    return int(product.get("totalQuantity") or 0)


def build_catalog_row(
    client: WildberriesPythonClient,
    product: dict[str, Any],
    card_info: dict[str, Any] | None,
) -> dict[str, Any]:
    groups = grouped_options(card_info)
    photo_count = int((card_info or {}).get("media", {}).get("photo_count") or product.get("pics") or 0)
    name = normalize_text(product.get("name") or (card_info or {}).get("imt_name"))
    rating = product.get("nmReviewRating") or product.get("reviewRating") or product.get("rating")
    review_count = int(product.get("nmFeedbacks") or product.get("feedbacks") or 0)
    country = find_characteristic(groups, "Страна производства")

    return {
        "product_url": build_product_url(int(product["id"])),
        "article": str(product["id"]),
        "name": name,
        "price": get_price(product),
        "description": normalize_text((card_info or {}).get("description")),
        "image_urls": ", ".join(client.build_image_urls(int(product["id"]), photo_count)),
        "characteristics": json.dumps(groups, ensure_ascii=False, indent=2),
        "seller_name": normalize_text(product.get("supplier")),
        "seller_url": build_seller_url(product.get("supplierId")),
        "sizes": ", ".join(get_sizes(product)),
        "stock": get_total_stock(product),
        "rating": float(rating) if rating is not None else None,
        "review_count": review_count,
        "country": country,
    }


def write_workbook(file_path: Path, rows: list[dict[str, Any]], sheet_name: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    headers = [
        ("Ссылка на товар", "product_url", 28),
        ("Артикул", "article", 14),
        ("Название", "name", 38),
        ("Цена", "price", 12),
        ("Описание", "description", 56),
        ("Ссылки на изображения", "image_urls", 72),
        ("Все характеристики", "characteristics", 60),
        ("Название селлера", "seller_name", 32),
        ("Ссылка на селлера", "seller_url", 28),
        ("Размеры товара", "sizes", 24),
        ("Остатки по товару", "stock", 18),
        ("Рейтинг", "rating", 12),
        ("Количество отзывов", "review_count", 18),
    ]

    worksheet.append([header for header, _, _ in headers])
    for row in rows:
        worksheet.append([row[key] for _, key, _ in headers])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:M{worksheet.max_row}"

    for column_index, (header, _, width) in enumerate(headers, start=1):
        column_letter = worksheet.cell(row=1, column=column_index).column_letter
        worksheet.column_dimensions[column_letter].width = width
        header_cell = worksheet.cell(row=1, column=column_index)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row_index in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row_index, column=1).style = "Hyperlink"
        worksheet.cell(row=row_index, column=1).hyperlink = worksheet.cell(row=row_index, column=1).value
        worksheet.cell(row=row_index, column=9).style = "Hyperlink"
        worksheet.cell(row=row_index, column=9).hyperlink = worksheet.cell(row=row_index, column=9).value
        worksheet.cell(row=row_index, column=4).number_format = "#,##0.00"

        for column_index in range(1, 14):
            wrap = column_index in {5, 6, 7, 8, 9, 10}
            worksheet.cell(row=row_index, column=column_index).alignment = Alignment(
                vertical="top",
                wrap_text=wrap,
            )

    workbook.save(file_path)


def main() -> None:
    args = parse_args()
    started_at = time.time()

    query = args.query
    dest = args.dest
    headless = parse_bool(args.headless)
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    print(
        f'[start] Запуск Python-выгрузки Wildberries по запросу "{query}" '
        f"(dest={dest}, headless={headless})."
    )

    client = WildberriesPythonClient(
        query=query,
        dest=dest,
        headless=headless,
        timeout_seconds=args.timeout_seconds,
    )

    try:
        client.start()
        product_ids = collect_product_ids(client, args.limit)
        print(f"[search] итоговое количество товаров к обработке: {len(product_ids)}.")

        detail_map = collect_detail_map(client, product_ids, args.batch_size)
        card_info_map = collect_card_info_map(client, product_ids, args.card_workers)

        rows: list[dict[str, Any]] = []
        for product_id in product_ids:
            product = detail_map.get(product_id)
            if product is None:
                continue
            rows.append(build_catalog_row(client, product, card_info_map.get(product_id)))

        filtered_rows = [
            row
            for row in rows
            if row["rating"] is not None
            and row["price"] is not None
            and float(row["rating"]) >= 4.5
            and float(row["price"]) <= 10000
            and "россия" in normalize_text(row["country"]).lower()
        ]

        full_catalog_path = output_dir / "wildberries_catalog_full.xlsx"
        filtered_catalog_path = output_dir / "wildberries_catalog_filtered.xlsx"

        write_workbook(full_catalog_path, rows, "Каталог")
        write_workbook(filtered_catalog_path, filtered_rows, "Фильтр")

        print(f"[done] Полный каталог: {full_catalog_path}")
        print(f"[done] Фильтрованный каталог: {filtered_catalog_path}")
        print(f"[done] Всего строк: {len(rows)}, после фильтра: {len(filtered_rows)}.")
        print(f"[done] Время выполнения: {format_duration(time.time() - started_at)}.")
    finally:
        client.close()


if __name__ == "__main__":
    main()
